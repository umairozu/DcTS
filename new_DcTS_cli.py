import math
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

R = 8.31446261815324 # gas const in J/(mol * k)
SEC_PER_WEEK = 7 * 24 * 3600
SEC_PER_YEAR = 365 * 24 * 3600

"""
We are Replicating 'Arrhenius Calculate' behavior using RawData.xlsx:
  -> k is obtained from linear fit of ln(C/Co) vs time in seconds:
        ln(c) = -k t + ln(c0) ### From Supp text S6
        
    (from the sheet):
  - D-DNA fit uses weeks 0,1,2 only 
  - E-DNA fit uses weeks 0,1,2,3
  
  - remaining fraction: C/C0 = exp(-k t)
  - half-life: ln(2)/k (converted to years)
  
  IMPORTANT Equation!!
  - Arrhenius: ln k = ln A - Ea/(R T)  (used for Fig 5G curve (Half life vs temp))
  - k is Decay rate constant
  - ea is activation energy
  
"""

@dataclass
class CassetteTapeDecay:

    k_d: Dict[int, float]  # D-DNA tape: temp_C -> k   (key, value)
    k_e: Dict[int, float]  # E-DNA tape: temp_C -> k   (key, value)
    ea_d: float = 90813.822     # Activation Energy decapsulated DNA
    ea_e: float = 133880.342    # Activation Energy encapsulated DNA
    lnA_d: float = None
    lnA_e: float = None

    blocks = {
        # format =
        # week no. : {temp: (start_row, end_row),...,...}
        0 : {60: (2,7), 65: (2,7), 70: (2,7)},
        1 : {60: (8, 13), 65: (14, 19), 70: (20,25)},
        2 : {60: (26, 31), 65: (32, 37), 70: (38, 43)},
        3 : {60: (44, 49), 65: (50, 55), 70: (56, 61)},

    }

    """Non-numeric junk removal (e.g N.A, blank cells, inf) for proper averaging,
    returns valid cells of type int ot float
    """
    @staticmethod
    def numeric_cells(file, col: str, r1: int, r2: int) -> List[float]:
        vals = []
        for r in range(r1, r2 + 1):
            v = file[f"{col}{r}"].value
            if isinstance(v,(int,float)) and not (math.isnan(v) or math.isinf(v)):
                vals.append(float(v))
        return vals

    """
    - ln_mean = C/ Co column
    - time_s = time in seconds
    ndarray is used instead of list bcz of speed and efficiency,
     ndarray is type hinting here, can't use array here!!!
     - Equation used -> ln (C / Co) = -kt
     - slope = -k
    """
    @staticmethod
    def fit_k(ln_means: np.ndarray, time_s: np.ndarray) -> float:
        slope, intercept = np.polyfit(time_s,ln_means,1)
        return float(-slope)

    """    
    - ln k = ln A - Ea/(R*T)  => ln A = ln k + Ea/(R*T)  
    -> A is some pre exponential factor in Arrhenius Eq.
    # due to limited data we are not fitting both Ea and lnA
    by linear Regression of lnK vs 1/T (wherein:
        slope m = -Ea / R --> Ea = -m / R
        intercept b = ln A), linear reg here would have yielded
    better results but since we only have 3 temps(60,65,75) and each
    k is little noisy, so fitting both parameters can be sensitive,
    as small change in k would change the slope and intercept a lot.
    
    Averaging ln A values is more stable approach in our case.
    """
    @staticmethod
    def fit_lnA(k_table: Dict[int, float], ea: float) -> float:
        vals = []
        for temp_C, k in k_table.items():
            T = 273.15 + float(temp_C)
            vals.append(math.log(k) + ea / (R * T))
        return float(sum(vals) / len(vals))

    #<------------------To be continued>#


    @staticmethod
    def from_rawdata_xlsx(xlsx_path: str) -> "CassetteTapeDecay":
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb["Arrhenius Calculate"]

        temps = [60, 65, 70]

        # D-DNA tape table data fetch from xlsx
        t_d = np.array([ws["H66"].value, ws["H67"].value, ws["H68"].value], dtype=float)
        cols_d = ["I", "J", "K"]
        k_d = {}
        for temp, col in zip(temps, cols_d):
            y = np.array([ws[f"{col}66"].value, ws[f"{col}67"].value, ws[f"{col}68"].value], dtype=float)
            slope, _ = np.polyfit(t_d, y, 1)
            k_d[temp] = float(-slope)

        # E-DNA tape table data fetch from xlsx
        t_e = np.array([ws["N66"].value, ws["N67"].value, ws["N68"].value, ws["N69"].value], dtype=float)
        cols_e = ["O", "P", "Q"]
        k_e = {}
        for temp, col in zip(temps, cols_e):
            y = np.array([ws[f"{col}66"].value, ws[f"{col}67"].value, ws[f"{col}68"].value, ws[f"{col}69"].value], dtype=float)
            slope, _ = np.polyfit(t_e, y, 1)
            k_e[temp] = float(-slope)

        model = CassetteTapeDecay(k_d=k_d, k_e=k_e)

        model.lnA_d = model.fit_lnA(model.k_d, model.ea_d)
        model.lnA_e = model.fit_lnA(model.k_e, model.ea_e)

        return model


    def k(self, temp_C: float, encapsulated: bool) -> float:
        temp_C_int = int(round(temp_C))
        if encapsulated and temp_C_int in self.k_e:
            return self.k_e[temp_C_int]
        if (not encapsulated) and temp_C_int in self.k_d:
            return self.k_d[temp_C_int]

        T = 273.15 + float(temp_C)
        if encapsulated:
            return math.exp(self.lnA_e - self.ea_e / (R * T))
        return math.exp(self.lnA_d - self.ea_d / (R * T))

    def remaining_fraction(self, temp_C: float, encapsulated: bool, weeks: float) -> float:
        t_sec = float(weeks) * SEC_PER_WEEK
        return math.exp(-self.k(temp_C, encapsulated) * t_sec)

    def half_life_years(self, temp_C: float, encapsulated: bool) -> float:
        k_val = self.k(temp_C, encapsulated)
        return (math.log(2) / k_val) / SEC_PER_YEAR

