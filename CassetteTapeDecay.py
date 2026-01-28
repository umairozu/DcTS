import math
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
from scipy import stats
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

R = 8.31446261815324 # gas const in J/(mol * k)
SEC_PER_WEEK = 7 * 24 * 3600
SEC_PER_YEAR = 365 * 24 * 3600

"""
We are Replicating 'Arrhenius Calculate' behavior using RawData.xlsx:
  -> k is obtained from linear fit of ln(C/Co) vs time in seconds:
        ln(c) = -k t + ln(c0) ### From Supp text S6
        
    (from the sheet):
  - D-DNA fit uses weeks 0,1,2 only 
  - E-DNA fit uses weeks 0,1,2,3
  
  - remaining fraction: C/C0 = exp(-k t)
  - half-life: ln(2)/k (converted to years)
  
  IMPORTANT Equation!!
  - Arrhenius: ln k = ln A - Ea/(R T)  (used for Fig 5G curve (Half life vs temp))
  - k is Decay rate constant
  - ea is activation energy
  
"""

@dataclass
class CassetteTapeDecay:

    k_dDNA: Dict[int, float]  # D-DNA tape: temp_C -> k   (key, value)
    k_eDNA: Dict[int, float]  # E-DNA tape: temp_C -> k   (key, value)
    ea_d: float = 90813.822     # Activation Energy decapsulated DNA
    ea_e: float = 133880.342    # Activation Energy encapsulated DNA
    lnA_dDNA: float = None
    lnA_eDNA: float = None

    blocks = {
        # format =
        # week no. : {temp: (start_row, end_row),...,...}
        0 : {60: (2,7), 65: (2,7), 70: (2,7)},
        1 : {60: (8, 13), 65: (14, 19), 70: (20,25)},
        2 : {60: (26, 31), 65: (32, 37), 70: (38, 43)},
        3 : {60: (44, 49), 65: (50, 55), 70: (56, 61)},

    }

    """Non-numeric junk removal (e.g N.A, blank cells, inf) for proper averaging,
    returns valid cells of type int ot float
    """
    @staticmethod
    def numeric_cells(file, col: str, r1: int, r2: int) -> List[float]:
        vals = []
        for r in range(r1, r2 + 1):
            v = file[f"{col}{r}"].value
            if isinstance(v,(int,float)) and not (math.isnan(v) or math.isinf(v)):
                vals.append(float(v))
        return vals

    """
    - ln_mean = C/ Co column
    - time_s = time in seconds
    ndarray is used instead of list bcz of speed and efficiency,
     ndarray is type hinting here, can't use array here!!!
     - Equation used -> ln (C / Co) = -kt
     - slope = -k
    """
    @staticmethod
    def fit_k(ln_means: np.ndarray, time_s: np.ndarray) -> float:
        slope, intercept = np.polyfit(time_s,ln_means,1)
        return float(-slope)

    """    
    - ln k = ln A - Ea/(R*T)  => ln A = ln k + Ea/(R*T)  
    -> A is some pre exponential factor in Arrhenius Eq.
    # due to limited data we are not fitting both Ea and lnA
    by linear Regression of lnK vs 1/T (wherein:
        slope m = -Ea / R --> Ea = -m / R
        intercept b = ln A), linear reg here would have yielded
    better results but since we only have 3 temps(60,65,75) and each
    k is little noisy, so fitting both parameters can be sensitive,
    as small change in k would change the slope and intercept a lot.
    
    Averaging ln A values is more stable approach in our case.
    """
    @staticmethod
    def fit_lnA(k_table: Dict[int, float], ea: float) -> float:
        vals = []
        for temp_C, k in k_table.items():
            T = 273.15 + float(temp_C)
            vals.append(math.log(k) + ea / (R * T))
        return float(sum(vals) / len(vals))

    """Retrieve data from xlsx for related calculations"""
    @classmethod
    def from_xlsx(cls, xlsx_path: str) -> "CassetteTapeDecay":
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb["Arrhenius Calculate"]

        ea_d = 10923*8.314 #hard coding from file
        ea_e = 16103*8.314 #hard coding from file

        # concentration columns for D-DNA & E-DNA
        col_D = "K"
        col_E = "Q"

        temps = [60,65,70]

        """Taking average of rows from ng/ul columns for that week"""
        def ln_conc_mean(col: str, week: int, temperature: int) -> Optional[float]:
            r1, r2 = cls.blocks[week][temperature]
            vals = cls.numeric_cells(ws,col,r1,r2)
            if len(vals) == 0:
                return None
            return float(math.log(np.mean(vals)))

        t0 = 0.0
        t1 = 1.0 * SEC_PER_WEEK
        t2 = 2.0 * SEC_PER_WEEK
        t3 = 3.0 * SEC_PER_WEEK

        # D-DNA
        k_dDNA : Dict[int, float] = {}
        for temp in temps:
            y0 = ln_conc_mean(col_D,0,temp)
            y1 = ln_conc_mean(col_D,1,temp)
            y2 = ln_conc_mean(col_D,2,temp)
            """ 
            y3 = No y3, it is most probably left intentionally
             maybe to be consistent with the fitting, as data for
             60 and 65 degrees is available but not for 70 degree
             in week 3
             """

            time = np.array([t0,t1,t2],dtype= float)
            ln_conc_meanS = np.array([y0,y1,y2],dtype=float)
            #k_dDNA[temp] = cls.fit_k(ln_conc_meanS,time)
            # ^^^^^^^^
            #  using all ln_conc of a particular temp calculate
            # one K value (e.g K @ 60 degree via ln_con(w0,w1,w2 etc @60 C))

            k_dDNA[temp] = cls.fit_k_lingress(ws, col="K", temp=temp, max_week=3)
            #^^^^^^^^^^^using this linear regression based k for now instead

        # E-DNA
        k_eDNA : Dict[int,float] = {}
        for temp in temps:
            y0 = ln_conc_mean(col_E, 0,temp)
            y1 = ln_conc_mean(col_E,1,temp)
            y2 = ln_conc_mean(col_E,2,temp)
            y3 = ln_conc_mean(col_E,3,temp)

            time = np.array([t0,t1,t2,t3], dtype= float)
            ln_conc_meanS = np.array([y0,y1,y2,y3,],dtype= float)

            #k_eDNA[temp] = cls.fit_k(ln_conc_meanS,time)
            k_eDNA[temp] = cls.fit_k_lingress(ws, col="Q", temp=temp, max_week=3)

        lnA_dDNA = cls.fit_lnA(k_dDNA,ea_d)
        lnA_eDNA = cls.fit_lnA(k_eDNA,ea_e)

        return cls(k_dDNA = k_dDNA, k_eDNA = k_eDNA,ea_d = ea_d, ea_e = ea_e, lnA_dDNA = lnA_dDNA, lnA_eDNA = lnA_eDNA)

    def k(self, temp_C: float, encapsulated: bool) -> float:
        temp = int(round(temp_C))
        """if encapsulated and temp in self.k_eDNA:
            return self.k_eDNA[temp]
        if (not encapsulated) and temp in self.k_dDNA:
            return self.k_dDNA[temp]"""

        """If your temp not in the dict, then convert the temp
        and put it into the ARRHENIUS EQUATION"""

        T = 273.15 + float(temp_C)
        if encapsulated:
            return math.exp(self.lnA_eDNA -(self.ea_e / (R * T)))
        return math.exp(self.lnA_dDNA - (self.ea_d / (R * T)))

    """
        - remaining_dna_frac = C / Co
        - formula used ->>>>   lnC - lnCo = -kt 
                               ln (C / Co) = -kt
                               C / Co = e^-kt
    """
    def remaining_dna_frac(self, temp_C: float, encapsulated: bool, week: float) -> float:
        t = SEC_PER_WEEK * week
        if (not encapsulated) and (temp_C >= 70.0) and (week > 2.0):
            return 1e-12 #kept it undetectable instead of 0.0 intentionally
        return math.exp(-self.k(temp_C,encapsulated) * t)

    """
    - half-life calculation
    - formula = ln(2) / k(temp)
        divide by Sec_per_year for half-life (Years) as in Fig 5.G
    """
    def half_life(self, temp_C: float, encapsulated: bool) -> float:
        return (math.log(2) / self.k(temp_C,encapsulated)) / SEC_PER_YEAR


    """
    calculating k via linear regression of ln (frac) vs time (t)
    where frac ---> C / Co 
    """
    @classmethod
    def fit_k_lingress(cls, ws, col: str, temp: int, max_week: int) -> float:

        c0_vals = cls.numeric_cells(ws, col, 2, 7)
        c0 = float(np.mean(c0_vals))

        xs: List[float] = []
        ys: List[float] = []

        for _v in c0_vals:
            xs.append(0.0)
            ys.append(0.0)

        for week in range(1, max_week + 1):
            r1, r2 = cls.blocks[week][temp]
            vals = cls.numeric_cells(ws, col, r1, r2)
            t = week * SEC_PER_WEEK
            for c in vals:
                frac = c / c0
                if frac > 0:
                    xs.append(t)
                    ys.append(math.log(frac))

        res =  stats.linregress(xs, ys)
        return float(-res.slope)























#<--------------------------------------EXTRA's-------------------------------------------------------------------------------->#


    """    def empirical_fractions(self, xlsx_path: str, encapsulated: bool) -> Dict[Tuple[int, int], List[float]]:
            """"""
            Returns replicate fractions (C/C0) per (tempC, week) computed from ng/uL columns,
            so it works even if Excel formula cache isn't present.
            """"""
            wb = load_workbook(xlsx_path, data_only=True)
            ws = wb["Arrhenius Calculate"]
            col = "Q" if encapsulated else "K"

            c0_vals = self.numeric_cells(ws, col, 2, 7)
            c0 = float(np.mean(c0_vals))

            out: Dict[Tuple[int, int], List[float]] = {}
            for week, mapping in self.blocks.items():
                for temp, (r1, r2) in mapping.items():
                    vals = self.numeric_cells(ws, col, r1, r2)
                    if week == 0:
                        out[(temp, week)] = [1.0]
                    else:
                        out[(temp, week)] = [v / c0 for v in vals] if vals else []
            return out
    """
